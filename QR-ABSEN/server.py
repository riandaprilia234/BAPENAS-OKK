import json
import os
import sys
import threading
from datetime import datetime, timezone, timedelta

WIB = timezone(timedelta(hours=7))
from flask import Flask, render_template, request, jsonify, Response, session, redirect

import requests
import io
import openpyxl

# Ensure current directory is in sys.path
sys.path.insert(0, os.path.dirname(__file__))

import supabase_db

template_dir = os.path.join(os.path.dirname(__file__), "templates")
app = Flask(__name__, template_folder=template_dir)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "BAPENAS2026_OKK_UNDIKA_SECRET_KEY_X9#mK")

DATA_FILE = os.path.join(os.path.dirname(__file__), "data_absen.json")
ADMIN_PIN = os.environ.get("ADMIN_PIN", "PanitiaOKK2026#")

data_lock = threading.Lock()

def load_data():
    if supabase_db.is_supabase_configured():
        sesi_dict, absen_dict = supabase_db.get_all_absen_grouped()
        return {"sesi": sesi_dict, "absen": absen_dict}
    
    with data_lock:
        if os.path.exists(DATA_FILE):
            try:
                with open(DATA_FILE, "r") as f:
                    return json.load(f)
            except Exception as e:
                print("Error reading local JSON:", e)
        return {"sesi": {}, "absen": {}}

def save_data_local(data):
    with data_lock:
        with open(DATA_FILE, "w") as f:
            json.dump(data, f, indent=4)

SCRIPT_URL = "https://script.google.com/macros/s/AKfycbwscj6QY55qJwWW2P8PJZSqIGliKDM_CpWnnYU8m-HB4JG_GI1p8xII-mHFpA1FzRkN/exec"

@app.route("/")
def index():
    if session.get("panitia"):
        return redirect("/panitia")
    return render_template("login.html")

@app.route("/panitia", methods=["GET", "POST"])
def panitia():
    base_url = request.host_url.rstrip("/")
    if request.method == "POST":
        if request.form.get("pin") == ADMIN_PIN:
            session["panitia"] = True
            data_absen = load_data()
            return render_template("panitia.html", sesi=data_absen["sesi"], base_url=base_url)
        return render_template("login.html", error="PIN Rahasia Salah! Silakan coba lagi.")
    
    if session.get("panitia"):
        data_absen = load_data()
        return render_template("panitia.html", sesi=data_absen["sesi"], base_url=base_url)
        
    return render_template("login.html")

@app.route("/panitia/buka-sesi", methods=["POST"])
def buka_sesi():
    nama = request.form.get("nama")
    if supabase_db.is_supabase_configured():
        sesi_id = supabase_db.create_sesi(nama)
    else:
        data_absen = load_data()
        sesi_id = str(len(data_absen["sesi"]) + 1)
        data_absen["sesi"][sesi_id] = {"nama": nama, "aktif": True}
        save_data_local(data_absen)
    return jsonify({"status": "ok", "sesi_id": sesi_id})

@app.route("/panitia/tutup-sesi/<sesi_id>", methods=["POST"])
def tutup_sesi(sesi_id):
    if supabase_db.is_supabase_configured():
        supabase_db.tutup_sesi(sesi_id)
    else:
        data_absen = load_data()
        if sesi_id in data_absen["sesi"]:
            data_absen["sesi"][sesi_id]["aktif"] = False
            save_data_local(data_absen)
    return jsonify({"status": "ok"})

@app.route("/tampilkan-qr/<sesi_id>")
def tampilkan_qr(sesi_id):
    data_absen = load_data()
    if sesi_id not in data_absen["sesi"]:
        return render_template("sesi_ditutup.html", sesi_id=None, sesi_nama="Sesi Tidak Ditemukan")
    
    sesi_info = data_absen["sesi"][sesi_id]
    if not sesi_info.get("aktif"):
        return render_template("sesi_ditutup.html", sesi_id=sesi_id, sesi_nama=sesi_info.get("nama"))

    base_url = request.host_url.rstrip("/")
    return render_template("tampilan_qr.html", sesi_id=sesi_id, sesi=sesi_info, base_url=base_url)

@app.route("/scan")
def scan():
    sesi_id = request.args.get("sesi")
    data_absen = load_data()
    if sesi_id not in data_absen["sesi"] or not data_absen["sesi"][sesi_id]["aktif"]:
        return render_template("scan_error.html", error="Sesi Ditutup", detail="Sesi ini sudah tidak menerima absen.")
    return render_template("form_absen.html", sesi_id=sesi_id)

@app.route("/submit-absen", methods=["POST"])
def submit_absen():
    sesi_id = request.form.get("sesi_id")
    nim = request.form.get("nim")
    nama = request.form.get("nama")
    agora = request.form.get("agora", "-")
    waktu = datetime.now(WIB).strftime("%Y-%m-%d %H:%M:%S")

    data_absen = load_data()

    if sesi_id not in data_absen["sesi"] or not data_absen["sesi"][sesi_id]["aktif"]:
        return render_template("scan_error.html", error="Sesi Ditutup", detail="Sesi ini sudah ditutup.")

    absen_sesi = data_absen["absen"].get(sesi_id, {})

    if nim in absen_sesi:
        return render_template("scan_sukses.html", sudah=True, nim=nim, nama=nama, sesi=data_absen["sesi"][sesi_id])

    if supabase_db.is_supabase_configured():
        supabase_db.submit_absen(sesi_id, nim, nama, agora, waktu)
    else:
        if sesi_id not in data_absen["absen"]:
            data_absen["absen"][sesi_id] = {}
        data_absen["absen"][sesi_id][nim] = {"nama": nama, "agora": agora, "waktu": waktu}
        save_data_local(data_absen)

    try:
        requests.post(SCRIPT_URL, data={"nim": nim, "nama": nama, "agora": agora, "sesi": data_absen["sesi"][sesi_id]["nama"], "waktu": waktu}, timeout=3)
    except:
        pass

    return render_template("scan_sukses.html", sudah=False, nim=nim, nama=nama, waktu=waktu, sesi=data_absen["sesi"][sesi_id])

@app.route("/download-excel")
def download_excel():
    if not session.get("panitia"):
        return render_template("login.html", error="Khusus Panitia! Masukkan PIN Rahasia terlebih dahulu.")
        
    data_absen = load_data()
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    data_per_agora = {}
    
    for sesi_id, sesi_info in data_absen["sesi"].items():
        sesi_nama = sesi_info["nama"]
        if sesi_id in data_absen["absen"]:
            for nim, maba in data_absen["absen"][sesi_id].items():
                agora = maba.get("agora", "Tidak Diketahui")
                if agora not in data_per_agora:
                    data_per_agora[agora] = []
                data_per_agora[agora].append([sesi_nama, nim, maba["nama"], maba["waktu"]])
                
    if not data_per_agora:
        ws = wb.create_sheet(title="Kosong")
        ws.append(["Belum ada data"])
    else:
        for agora, rows in data_per_agora.items():
            safe_title = str(agora)[:31].replace("/", "-").replace("\\", "-")
            ws = wb.create_sheet(title=safe_title)
            ws.append(["Nama Sesi", "NIM", "Nama Lengkap", "Waktu Absen"])
            for r in rows:
                ws.append(r)
                
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment;filename=Rekap_MultiTab_OKK_ETHIVATION.xlsx"}
    )

@app.route("/rekap/<sesi_id>")
def rekap_sesi(sesi_id):
    if not session.get("panitia"):
        return render_template("login.html", error="Khusus Panitia! Masukkan PIN Rahasia terlebih dahulu.")
        
    data_absen = load_data()
    if sesi_id not in data_absen["sesi"]:
        return "Sesi tidak ditemukan.", 404
    
    sesi = data_absen["sesi"][sesi_id]
    sesi["id"] = sesi_id
    absen_dict = data_absen["absen"].get(sesi_id, {})
    
    absen_list = []
    for nim, info in absen_dict.items():
        absen_list.append({
            "nim": nim,
            "nama": info.get("nama", ""),
            "agora": info.get("agora", "-"),
            "waktu": info.get("waktu", "")
        })
        
    return render_template("rekap.html", sesi=sesi, absen_list=absen_list, base_url=request.host_url.rstrip("/"))

@app.route("/rekap")
def rekap_latest():
    if not session.get("panitia"):
        return render_template("login.html", error="Khusus Panitia! Masukkan PIN Rahasia terlebih dahulu.")
        
    data_absen = load_data()
    if not data_absen["sesi"]:
        return "Belum ada sesi.", 404
    latest_id = list(data_absen["sesi"].keys())[-1]
    return rekap_sesi(latest_id)

@app.route("/api/recent-checkins")
def api_recent_checkins():
    data_absen = load_data()
    semua_checkins = []
    
    for sesi_id, absen_dict in data_absen.get("absen", {}).items():
        for nim, info in absen_dict.items():
            semua_checkins.append({
                "nim": nim,
                "nama": info.get("nama", ""),
                "agora": info.get("agora", "-"),
                "waktu": info.get("waktu", "")
            })
            
    semua_checkins.sort(key=lambda x: x["waktu"], reverse=True)
    return jsonify(semua_checkins[:10])

@app.route("/api/live-absen/<sesi_id>")
def api_live_absen(sesi_id):
    data_absen = load_data()
    if sesi_id not in data_absen["absen"]:
        return jsonify([])
    
    semua_absen = list(data_absen["absen"][sesi_id].values())
    terbaru = semua_absen[-15:]
    terbaru.reverse()
    
    return jsonify(terbaru)

if __name__ == "__main__":
    from waitress import serve
    PORT = int(os.environ.get("PORT", 5000))
    print(f"  QR ABSEN DINAMIS - OKK ETHIVATION")
    print(f"  Berjalan di port {PORT} dengan Waitress...")
    serve(app, host="0.0.0.0", port=PORT, threads=32)
