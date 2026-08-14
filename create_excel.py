import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Rekap Teguran Mentor'

headers = [
    'Waktu (Jam Kejadian)',
    'Nama Peserta & Kelompok',
    'Jenis Pelanggaran Ringan',
    'Tindakan',
    'Nama Mentor Penindak',
    'Keterangan / Alasan'
]
ws.append(headers)

header_fill = PatternFill(start_color='FF1F4E78', end_color='FF1F4E78', fill_type='solid')
header_font = Font(color='FFFFFFFF', bold=True)
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

for col_num, header_title in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    ws.column_dimensions[get_column_letter(col_num)].width = 30

dv = DataValidation(type='list', formula1='"Teguran 1,Teguran 2,Proses Sanksi"', allow_blank=True)
dv.error ='Harap pilih dari dropdown'
dv.errorTitle = 'Pilihan Tidak Valid'
dv.prompt = 'Pilih tindakan'
dv.promptTitle = 'Pilih Tindakan'

ws.add_data_validation(dv)
dv.add('D2:D1000')

wb.save(r'C:\Users\LENOVO\Documents\BAPENAS 2026\FORMAT TEGURAN MENTOR OKK 2026.xlsx')
print('Excel file created successfully.')
